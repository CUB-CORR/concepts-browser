"""Tabular export of the concepts list (CSV / XLSX) for admins.

One row per active name in the export taxonomy — the same rows the list endpoint returns — one
name column per taxonomy, plus the concept-level fields (concept id, description, group size,
version, sources, types, clinical documentation, PICO frame, study team). `include_configs`
appends the latest
published `json`/`py` per source as extra columns; those payloads are deliberately not read
otherwise (see `services.latest_published_index`).
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime

from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session

from . import services
from .models import RELATIONSHIP_ALIAS, Concept, ConceptTaxonomy, Config, Source, Taxonomy, _utcnow

CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _config_payload_index(
    db: Session, concept_ids: list[int], at_time: datetime | None
) -> dict[tuple[int, int], tuple[dict, str | None]]:
    """{(concept_id, source_id): (json_def, python_code)} for the latest published config,
    same selection rule as `services.latest_published_index` but carrying the payloads."""
    filters = [
        Config.status == "published",
        Config.version_no.isnot(None),
        Config.concept_id.in_(concept_ids),
    ]
    if at_time is not None:
        filters.append(Config.created_at <= at_time)

    mv = (
        select(
            Config.concept_id,
            Config.source_id,
            func.max(Config.version_no).label("mv"),
        )
        .where(*filters)
        .group_by(Config.concept_id, Config.source_id)
        .subquery()
    )
    rows = db.execute(
        select(Config.concept_id, Config.source_id, Config.json_def, Config.python_code)
        .join(
            mv,
            and_(
                Config.concept_id == mv.c.concept_id,
                Config.source_id == mv.c.source_id,
                Config.version_no == mv.c.mv,
            ),
        )
        .where(Config.status == "published")
    )
    return {(r.concept_id, r.source_id): (r.json_def, r.python_code) for r in rows}


def build_export_table(
    db: Session,
    tax: Taxonomy,
    *,
    ids: list[int] | None,
    at_time: datetime | None,
    include_configs: bool,
) -> tuple[list[str], list[list[str | int | None]]]:
    """Assemble (headers, rows) for every name registered in `tax` (optionally only `ids`)."""
    # Export taxonomy first, the rest in key order — same names the app shows.
    taxonomies = sorted(db.scalars(select(Taxonomy)), key=lambda t: (t.key != tax.key, t.key))
    source_keys = {s.id: s.key for s in db.scalars(select(Source))}

    now = at_time or _utcnow()
    # One row per active pointer, matching the list endpoint: an alias exports as its own row,
    # and each member of a group exports separately (told apart by `concept_id`).
    concept_filter = [
        ConceptTaxonomy.taxonomy_id == tax.id,
        services.pointer_active_at(now),
    ]
    if ids is not None:
        concept_filter.append(Concept.id.in_(ids))
    rows = db.execute(
        select(Concept, ConceptTaxonomy)
        .join(ConceptTaxonomy, ConceptTaxonomy.concept_id == Concept.id)
        .where(*concept_filter)
        .order_by(ConceptTaxonomy.identifier, Concept.id)
    ).all()
    concept_ids = [c.id for c, _ in rows]
    groups = services.group_size_index(db, tax.id, now)

    # All taxonomy names of the exported concepts: (concept_id, taxonomy_id) -> identifier.
    # A retired name or an alias only fills a cell no active primary name claims.
    names: dict[tuple[int, int], str] = {}
    for ct in db.scalars(
        select(ConceptTaxonomy)
        .where(ConceptTaxonomy.concept_id.in_(concept_ids))
        .order_by(
            ConceptTaxonomy.deprecated_at.is_not(None),
            func.coalesce(ConceptTaxonomy.relationship, "") == RELATIONSHIP_ALIAS,
            ConceptTaxonomy.identifier,
        )
    ):
        names.setdefault((ct.concept_id, ct.taxonomy_id), ct.identifier)

    index = services.latest_published_index(db, at_time=at_time)
    payloads = (
        _config_payload_index(db, concept_ids, at_time) if include_configs else {}
    )
    # Config columns only for sources that actually contribute to the exported rows.
    config_source_ids = sorted(
        {sid for cid, sid in payloads if cid in set(concept_ids)},
        key=lambda sid: source_keys.get(sid, ""),
    )

    headers = (
        ["concept_id"]
        + [f"name ({t.key})" for t in taxonomies]
        + [
            "display_name",
            "description",
            "group_size",
            "version",
            "sources",
            "types",
            "doc_clinical",
            "doc_implementation",
            "doc_caveats",
            "doc_status",
            "notion_url",
        ]
        + [
            col
            for sid in config_source_ids
            for col in (f"config json ({source_keys[sid]})", f"config py ({source_keys[sid]})")
        ]
    )

    out: list[list[str | int | None]] = []
    for concept, ct in rows:
        configs = index.get(concept.id, [])
        row: list[str | int | None] = [concept.id]
        # The export taxonomy's cell is this row's own identifier — the row *is* that name.
        row += [
            ct.identifier if t.id == tax.id else names.get((concept.id, t.id))
            for t in taxonomies
        ]
        row += [
            ct.display_name,
            concept.description,
            groups.get(ct.identifier, 1),
            max((c.version_no for c in configs), default=None),
            ", ".join(
                sorted({source_keys[c.source_id] for c in configs if c.source_id in source_keys})
            ),
            ", ".join(sorted({c.type for c in configs})),
            concept.doc_clinical,
            concept.doc_implementation,
            concept.doc_caveats,
            concept.doc_status,
            concept.notion_url,
        ]
        for sid in config_source_ids:
            payload = payloads.get((concept.id, sid))
            row += [
                json.dumps(payload[0], indent=1, ensure_ascii=False) if payload else None,
                payload[1] if payload else None,
            ]
        out.append(row)
    return headers, out


def to_csv(headers: list[str], rows: list[list[str | int | None]]) -> bytes:
    """UTF-8 with BOM so Excel detects the encoding when double-clicked."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx(headers: list[str], rows: list[list[str | int | None]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Concepts"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(row)
        # openpyxl treats a leading "=" as a formula; force those cells back to text so
        # documentation prose can never execute as a spreadsheet formula.
        for cell in ws[ws.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"

    for col_no, header in enumerate(headers, start=1):
        width = max(
            len(header),
            max((len(str(r[col_no - 1] or "").split("\n", 1)[0]) for r in rows), default=0),
        )
        ws.column_dimensions[get_column_letter(col_no)].width = min(width + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
