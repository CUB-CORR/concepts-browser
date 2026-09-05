"""Smoke test for the admin concepts export (CSV/XLSX). Run: uv run python tests/export_smoke.py"""
import csv
import io
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

db_path = os.path.join(tempfile.mkdtemp(), "export_smoke.db")
os.environ["DATABASE_URL"] = f"sqlite:///{db_path}"
os.environ["JWT_SECRET"] = "test-secret"
os.environ["APP_SHARED_SECRET"] = ""
os.environ["BOOTSTRAP_ADMIN_USERNAME"] = "admin"
os.environ["BOOTSTRAP_ADMIN_PASSWORD"] = "admin"
os.environ.setdefault("SCHEMA_DIR", "reference/schema")
os.environ["IMPORT_REFERENCE_ON_FIRST_RUN"] = "false"

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from api.main import app  # noqa: E402

ND_Z49 = {"type": "native_dynamic", "table_name": "diagnoses",
          "where_clause": "code LIKE 'Z49%'"}


def parse_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


with TestClient(app) as client:
    r = client.post("/auth/login", json={"username": "admin", "password": "admin"})
    assert r.status_code == 200, r.text
    h = {"authorization": f"Bearer {r.json()['access_token']}"}

    # --- fixture: two concepts; alpha gets a published config + documentation -----------
    r = client.post("/concepts", headers=h, json={
        "name": "alpha_var", "display_name": "Alpha", "description": "=SUM(A1) injected?",
    })
    assert r.status_code == 201, r.text
    alpha_id = r.json()["id"]
    r = client.post("/concepts", headers=h, json={"name": "beta_var"})
    assert r.status_code == 201, r.text
    beta_id = r.json()["id"]

    r = client.post("/concept/corr_v1/alpha_var/drafts", headers=h, json={
        "source": "cub_hdp", "empty": True, "type": "native_dynamic", "json": ND_Z49,
        "py": "def alpha_var(var, cohort): ...",
    })
    assert r.status_code == 201, r.text
    r = client.post(
        f"/concept/corr_v1/alpha_var/drafts/{r.json()['id']}/publish", headers=h, json={}
    )
    assert r.status_code == 200, r.text

    r = client.patch("/concept/corr_v1/alpha_var/documentation", headers=h, json={
        "doc_clinical": "Used for ICU scoring.", "doc_status": "In Production",
    })
    assert r.status_code == 200, r.text

    # --- CSV: full export --------------------------------------------------------------
    r = client.post("/concepts/export", headers=h, json={"format": "csv"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    assert 'filename="concepts_corr_v1_' in r.headers["content-disposition"]
    rows = {row["name (corr_v1)"]: row for row in parse_csv(r.content)}
    assert set(rows) == {"alpha_var", "beta_var"}, rows.keys()
    alpha = rows["alpha_var"]
    assert alpha["display_name"] == "Alpha" and alpha["version"] == "1"
    assert alpha["concept_id"] == str(alpha_id) and alpha["group_size"] == "1", alpha
    assert alpha["sources"] == "cub_hdp" and alpha["types"] == "native_dynamic"
    assert alpha["doc_clinical"] == "Used for ICU scoring."
    assert alpha["doc_status"] == "In Production"
    assert rows["beta_var"]["version"] == "" and rows["beta_var"]["sources"] == ""
    # config columns are absent unless asked for
    assert "config json (cub_hdp)" not in alpha

    # --- CSV: filtered ids (the app sends its filtered list) ----------------------------
    r = client.post("/concepts/export", headers=h, json={"format": "csv", "ids": [beta_id]})
    assert [row["name (corr_v1)"] for row in parse_csv(r.content)] == ["beta_var"]

    # --- CSV: include_configs adds the latest published json/py per source --------------
    r = client.post("/concepts/export", headers=h,
                    json={"format": "csv", "include_configs": True})
    rows = {row["name (corr_v1)"]: row for row in parse_csv(r.content)}
    assert "native_dynamic" in rows["alpha_var"]["config json (cub_hdp)"]
    assert rows["alpha_var"]["config py (cub_hdp)"].startswith("def alpha_var")
    assert rows["beta_var"]["config json (cub_hdp)"] == ""

    # --- XLSX --------------------------------------------------------------------------
    r = client.post("/concepts/export", headers=h, json={"format": "xlsx"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    ws = load_workbook(io.BytesIO(r.content)).active
    header = [c.value for c in ws[1]]
    assert header[:2] == ["concept_id", "name (corr_v1)"] and "doc_clinical" in header
    data = {row[1]: dict(zip(header, row)) for row in ws.iter_rows(min_row=2, values_only=True)}
    assert set(data) == {"alpha_var", "beta_var"}
    assert data["alpha_var"]["version"] == 1
    # a value starting with "=" must come back as text, not as a formula
    assert data["alpha_var"]["description"] == "=SUM(A1) injected?"

    # --- gates ---------------------------------------------------------------------------
    # admin only: a key narrowed to can_read may read concepts but not export them
    r = client.post("/api-keys", headers=h, json={"name": "reader", "scopes": ["can_read"]})
    assert r.status_code == 201, r.text
    key_h = {"authorization": f"Bearer {r.json()['key']}"}
    assert client.get("/concepts?project=internal", headers=key_h).status_code == 200
    assert client.post("/concepts/export", headers=key_h, json={}).status_code == 403
    # unauthenticated
    assert client.post("/concepts/export", json={}).status_code in (401, 403)
    # unknown taxonomy / bad format
    assert client.post("/concepts/export", headers=h,
                       json={"taxonomy": "nope"}).status_code == 404
    assert client.post("/concepts/export", headers=h,
                       json={"format": "pdf"}).status_code == 422

print("export smoke test passed")
